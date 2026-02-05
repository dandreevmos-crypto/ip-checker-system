# -*- coding: utf-8 -*-
"""
Модуль экспорта результатов проверки в различные форматы
Excel, CSV, HTML, JSON
"""

import os
import json
from typing import List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR, TrafficLightStatus
from models import ProductItem, CheckSession, CheckReport, RiskLevel
from risk_evaluator import RiskAssessment, TrafficLightReportGenerator


class ExportManager:
    """Менеджер экспорта результатов проверки"""

    # Цвета для статусов в Excel
    EXCEL_COLORS = {
        "red": "FF4444",
        "yellow": "FFBB33",
        "green": "00C851"
    }

    def __init__(self, output_dir: str = None):
        self.output_dir = Path(output_dir) if output_dir else OUTPUT_DIR
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_to_excel(self, session: CheckSession,
                         assessments: Dict[str, RiskAssessment],
                         filename: str = None) -> str:
        """
        Экспорт результатов в Excel с форматированием

        Args:
            session: Сессия проверки
            assessments: Словарь оценок {артикул: оценка}
            filename: Имя файла (опционально)

        Returns:
            Путь к созданному файлу
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ip_check_results_{session.session_id}_{timestamp}.xlsx"

        filepath = self.output_dir / filename

        # Создаем рабочую книгу
        wb = Workbook()

        # Лист с общей сводкой
        self._create_summary_sheet(wb, session, assessments)

        # Лист с детальными результатами
        self._create_details_sheet(wb, session, assessments)

        # Лист с рекомендациями
        self._create_recommendations_sheet(wb, session, assessments)

        # Лист со статистикой
        self._create_statistics_sheet(wb, session, assessments)

        # Удаляем дефолтный лист, если он пустой
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

        wb.save(filepath)
        return str(filepath)

    def _create_summary_sheet(self, wb: Workbook, session: CheckSession,
                               assessments: Dict[str, RiskAssessment]):
        """Создание листа со сводкой"""
        ws = wb.create_sheet("Сводка", 0)

        # Заголовки
        headers = [
            "Статус", "Артикул", "Название", "Оценка риска",
            "Критических", "Требуют внимания", "Источник", "Краткое описание"
        ]

        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Записываем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Записываем данные
        row = 2
        for item in session.items:
            assessment = assessments.get(item.article)
            if assessment is None:
                continue

            # Получаем цвет для статуса
            status_color = self.EXCEL_COLORS.get(assessment.overall_status.value, "FFFFFF")
            status_fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")

            # Иконка статуса
            status_icons = {"red": "🔴", "yellow": "🟡", "green": "🟢"}
            status_icon = status_icons.get(assessment.overall_status.value, "⚪")

            # Данные строки
            row_data = [
                status_icon,
                item.article,
                item.name,
                f"{assessment.overall_score:.0f}%",
                sum(1 for f in assessment.factors if f.severity == RiskLevel.RED),
                sum(1 for f in assessment.factors if f.severity == RiskLevel.YELLOW),
                item.image_source.source_type if item.image_source else "неизвестно",
                assessment.summary[:100] + "..." if len(assessment.summary) > 100 else assessment.summary
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border

                # Подсветка статуса
                if col == 1:
                    cell.fill = status_fill
                    cell.alignment = Alignment(horizontal='center')

            row += 1

        # Автоподбор ширины колонок
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Фиксация заголовка
        ws.freeze_panes = "A2"

    def _create_details_sheet(self, wb: Workbook, session: CheckSession,
                               assessments: Dict[str, RiskAssessment]):
        """Создание листа с детальными результатами"""
        ws = wb.create_sheet("Детали проверки")

        headers = [
            "Артикул", "Название", "Статус", "Категория фактора",
            "Название фактора", "Серьезность", "Описание"
        ]

        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        row = 2
        for item in session.items:
            assessment = assessments.get(item.article)
            if assessment is None or not assessment.factors:
                continue

            for factor in assessment.factors:
                severity_labels = {"red": "Критический", "yellow": "Предупреждение", "green": "OK"}
                severity_label = severity_labels.get(factor.severity.value, "Неизвестно")

                row_data = [
                    item.article,
                    item.name,
                    TrafficLightStatus.LABELS.get(assessment.overall_status.value, ""),
                    factor.category,
                    factor.name,
                    severity_label,
                    factor.description
                ]

                severity_color = self.EXCEL_COLORS.get(factor.severity.value, "FFFFFF")
                severity_fill = PatternFill(start_color=severity_color, end_color=severity_color, fill_type="solid")

                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if col == 6:  # Колонка серьезности
                        cell.fill = severity_fill

                row += 1

        # Автоподбор ширины
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        ws.freeze_panes = "A2"

    def _create_recommendations_sheet(self, wb: Workbook, session: CheckSession,
                                       assessments: Dict[str, RiskAssessment]):
        """Создание листа с рекомендациями"""
        ws = wb.create_sheet("Рекомендации")

        headers = ["Артикул", "Название", "Статус", "Рекомендации", "Требуется ручная проверка"]

        # Стили заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        row = 2
        for item in session.items:
            assessment = assessments.get(item.article)
            if assessment is None:
                continue

            recommendations_text = "\n".join(f"• {r}" for r in assessment.recommendations)
            manual_check_text = "\n".join(f"• {m}" for m in assessment.manual_check_items) if assessment.requires_manual_check else "Не требуется"

            status_icons = {"red": "🔴 ЗАПРЕЩЕНО", "yellow": "🟡 ВНИМАНИЕ", "green": "🟢 РАЗРЕШЕНО"}
            status_text = status_icons.get(assessment.overall_status.value, "")

            row_data = [
                item.article,
                item.name,
                status_text,
                recommendations_text,
                manual_check_text
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')

            row += 1

        # Ширина колонок
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 40

        ws.freeze_panes = "A2"

    def _create_statistics_sheet(self, wb: Workbook, session: CheckSession,
                                  assessments: Dict[str, RiskAssessment]):
        """Создание листа со статистикой"""
        ws = wb.create_sheet("Статистика")

        # Подсчет статистики
        total = len(session.items)
        red_count = sum(1 for a in assessments.values() if a.overall_status == RiskLevel.RED)
        yellow_count = sum(1 for a in assessments.values() if a.overall_status == RiskLevel.YELLOW)
        green_count = sum(1 for a in assessments.values() if a.overall_status == RiskLevel.GREEN)

        # Заголовок
        ws['A1'] = "Статистика проверки"
        ws['A1'].font = Font(bold=True, size=14)

        # Информация о сессии
        ws['A3'] = "ID сессии:"
        ws['B3'] = session.session_id
        ws['A4'] = "Дата проверки:"
        ws['B4'] = datetime.now().strftime("%d.%m.%Y %H:%M")
        ws['A5'] = "Всего товаров:"
        ws['B5'] = total

        # Статистика по статусам
        ws['A7'] = "Статистика по статусам:"
        ws['A7'].font = Font(bold=True)

        stats = [
            ("🔴 Запрещено (красный)", red_count, red_count/total*100 if total > 0 else 0, "FF4444"),
            ("🟡 Внимание (желтый)", yellow_count, yellow_count/total*100 if total > 0 else 0, "FFBB33"),
            ("🟢 Разрешено (зеленый)", green_count, green_count/total*100 if total > 0 else 0, "00C851")
        ]

        row = 8
        for label, count, percent, color in stats:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{percent:.1f}%")
            ws.cell(row=row, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            row += 1

        # Ширина колонок
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

    def export_to_csv(self, session: CheckSession,
                       assessments: Dict[str, RiskAssessment],
                       filename: str = None) -> str:
        """Экспорт в CSV"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ip_check_results_{session.session_id}_{timestamp}.csv"

        filepath = self.output_dir / filename

        # Подготовка данных
        data = []
        for item in session.items:
            assessment = assessments.get(item.article)
            if assessment is None:
                continue

            data.append({
                "Артикул": item.article,
                "Название": item.name,
                "Статус": TrafficLightStatus.LABELS.get(assessment.overall_status.value, ""),
                "Оценка риска": f"{assessment.overall_score:.0f}",
                "Критических факторов": sum(1 for f in assessment.factors if f.severity == RiskLevel.RED),
                "Предупреждений": sum(1 for f in assessment.factors if f.severity == RiskLevel.YELLOW),
                "Источник": item.image_source.source_type if item.image_source else "",
                "Рекомендации": "; ".join(assessment.recommendations),
                "Дата проверки": datetime.now().strftime("%d.%m.%Y %H:%M")
            })

        df = pd.DataFrame(data)
        df.to_csv(filepath, index=False, encoding='utf-8-sig')

        return str(filepath)

    def export_to_json(self, session: CheckSession,
                        assessments: Dict[str, RiskAssessment],
                        filename: str = None) -> str:
        """Экспорт в JSON"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ip_check_results_{session.session_id}_{timestamp}.json"

        filepath = self.output_dir / filename

        # Формируем структуру данных
        export_data = {
            "session": {
                "id": session.session_id,
                "created_at": session.created_at.isoformat(),
                "total_items": session.total_items,
                "statistics": {
                    "red": sum(1 for a in assessments.values() if a.overall_status == RiskLevel.RED),
                    "yellow": sum(1 for a in assessments.values() if a.overall_status == RiskLevel.YELLOW),
                    "green": sum(1 for a in assessments.values() if a.overall_status == RiskLevel.GREEN)
                }
            },
            "items": []
        }

        for item in session.items:
            assessment = assessments.get(item.article)
            if assessment is None:
                continue

            item_data = TrafficLightReportGenerator.format_assessment_for_export(assessment, item)
            export_data["items"].append(item_data)

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, ensure_ascii=False, indent=2)

        return str(filepath)

    def export_to_html(self, session: CheckSession,
                        assessments: Dict[str, RiskAssessment],
                        filename: str = None) -> str:
        """Экспорт в HTML для просмотра в браузере"""
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ip_check_results_{session.session_id}_{timestamp}.html"

        filepath = self.output_dir / filename

        # Подсчет статистики
        total = len(session.items)
        red_count = sum(1 for a in assessments.values() if a.overall_status == RiskLevel.RED)
        yellow_count = sum(1 for a in assessments.values() if a.overall_status == RiskLevel.YELLOW)
        green_count = sum(1 for a in assessments.values() if a.overall_status == RiskLevel.GREEN)

        html_content = f'''<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Результаты проверки ИС - {session.session_id}</title>
    <style>
        * {{ box-sizing: border-box; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            margin: 0;
            padding: 20px;
            background: #f5f5f5;
        }}
        .container {{ max-width: 1400px; margin: 0 auto; }}
        .header {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            margin-bottom: 20px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        h1 {{ margin: 0 0 10px 0; color: #333; }}
        .session-info {{ color: #666; font-size: 14px; }}
        .stats {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            margin-bottom: 20px;
        }}
        .stat-card {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            text-align: center;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .stat-card.red {{ border-left: 4px solid #FF4444; }}
        .stat-card.yellow {{ border-left: 4px solid #FFBB33; }}
        .stat-card.green {{ border-left: 4px solid #00C851; }}
        .stat-number {{ font-size: 36px; font-weight: bold; }}
        .stat-label {{ color: #666; margin-top: 5px; }}
        .results-table {{
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        table {{ width: 100%; border-collapse: collapse; }}
        th {{
            background: #4472C4;
            color: white;
            padding: 15px;
            text-align: left;
        }}
        td {{ padding: 15px; border-bottom: 1px solid #eee; }}
        tr:hover {{ background: #f9f9f9; }}
        .status-badge {{
            display: inline-block;
            padding: 5px 15px;
            border-radius: 20px;
            font-weight: bold;
            font-size: 12px;
        }}
        .status-red {{ background: #FF4444; color: white; }}
        .status-yellow {{ background: #FFBB33; color: #333; }}
        .status-green {{ background: #00C851; color: white; }}
        .score {{ font-weight: bold; }}
        .factors {{ font-size: 12px; color: #666; }}
        .recommendations {{
            max-width: 300px;
            font-size: 12px;
            color: #555;
        }}
        .recommendations ul {{
            margin: 0;
            padding-left: 15px;
        }}
        .legend {{
            background: white;
            padding: 15px;
            border-radius: 8px;
            margin-bottom: 20px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .legend h3 {{ margin: 0 0 10px 0; }}
        .legend-item {{ display: flex; align-items: center; margin: 5px 0; }}
        .legend-color {{
            width: 20px;
            height: 20px;
            border-radius: 50%;
            margin-right: 10px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Результаты проверки интеллектуальной собственности</h1>
            <div class="session-info">
                Сессия: {session.session_id} |
                Дата: {datetime.now().strftime("%d.%m.%Y %H:%M")} |
                Всего товаров: {total}
            </div>
        </div>

        <div class="stats">
            <div class="stat-card red">
                <div class="stat-number" style="color: #FF4444;">{red_count}</div>
                <div class="stat-label">Запрещено</div>
            </div>
            <div class="stat-card yellow">
                <div class="stat-number" style="color: #FFBB33;">{yellow_count}</div>
                <div class="stat-label">Требует проверки</div>
            </div>
            <div class="stat-card green">
                <div class="stat-number" style="color: #00C851;">{green_count}</div>
                <div class="stat-label">Разрешено</div>
            </div>
        </div>

        <div class="legend">
            <h3>Легенда статусов:</h3>
            <div class="legend-item">
                <div class="legend-color" style="background: #FF4444;"></div>
                <span><strong>КРАСНЫЙ</strong> - Использование запрещено. Высокий риск претензий.</span>
            </div>
            <div class="legend-item">
                <div class="legend-color" style="background: #FFBB33;"></div>
                <span><strong>ЖЕЛТЫЙ</strong> - Требуется дополнительная проверка.</span>
            </div>
            <div class="legend-item">
                <div class="legend-color" style="background: #00C851;"></div>
                <span><strong>ЗЕЛЕНЫЙ</strong> - Можно использовать. Рисков не выявлено.</span>
            </div>
        </div>

        <div class="results-table">
            <table>
                <thead>
                    <tr>
                        <th>Статус</th>
                        <th>Артикул</th>
                        <th>Название</th>
                        <th>Оценка</th>
                        <th>Факторы</th>
                        <th>Рекомендации</th>
                    </tr>
                </thead>
                <tbody>
'''

        # Добавляем строки таблицы
        for item in session.items:
            assessment = assessments.get(item.article)
            if assessment is None:
                continue

            status_class = f"status-{assessment.overall_status.value}"
            status_label = {
                "red": "ЗАПРЕЩЕНО",
                "yellow": "ВНИМАНИЕ",
                "green": "РАЗРЕШЕНО"
            }.get(assessment.overall_status.value, "")

            red_factors = sum(1 for f in assessment.factors if f.severity == RiskLevel.RED)
            yellow_factors = sum(1 for f in assessment.factors if f.severity == RiskLevel.YELLOW)

            recommendations_html = "<ul>" + "".join(f"<li>{r}</li>" for r in assessment.recommendations[:3]) + "</ul>"

            html_content += f'''
                    <tr>
                        <td><span class="status-badge {status_class}">{status_label}</span></td>
                        <td>{item.article}</td>
                        <td>{item.name}</td>
                        <td class="score">{assessment.overall_score:.0f}%</td>
                        <td class="factors">
                            🔴 {red_factors} критических<br>
                            🟡 {yellow_factors} предупреждений
                        </td>
                        <td class="recommendations">{recommendations_html}</td>
                    </tr>
'''

        html_content += '''
                </tbody>
            </table>
        </div>
    </div>
</body>
</html>
'''

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)

        return str(filepath)


if __name__ == "__main__":
    print("Модуль экспорта готов к использованию")
    print(f"Директория для экспорта: {OUTPUT_DIR}")
