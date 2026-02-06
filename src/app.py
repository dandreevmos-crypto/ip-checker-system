# -*- coding: utf-8 -*-
"""
Веб-приложение для проверки интеллектуальной собственности
Flask-based веб-интерфейс для менеджеров
"""

import os
import sys
import uuid
import json
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any

from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for
from flask_cors import CORS
from werkzeug.utils import secure_filename

# Добавляем путь к src
sys.path.insert(0, str(Path(__file__).parent))

from config import APP_CONFIG, DATA_DIR, OUTPUT_DIR, TRADEMARK_RESOURCES, IMAGE_SEARCH_RESOURCES, MKTU_CLASSES
from models import ProductItem, CheckSession, ImageSource, RiskLevel
from data_loader import DataLoader, TemplateGenerator
from trademark_checker import ComprehensiveTrademarkChecker
from image_checker import ComprehensiveImageChecker
from image_search_api import ComprehensiveImageSearcher
from risk_evaluator import RiskEvaluator, RiskAssessment
from export_manager import ExportManager
from database import (
    save_name_check, save_image_check,
    get_name_checks, get_image_checks,
    get_name_check_by_id, get_image_check_by_id,
    get_statistics, delete_check, clear_history
)

# Инициализация Flask
app = Flask(__name__,
            template_folder=str(Path(__file__).parent.parent / 'templates'),
            static_folder=str(Path(__file__).parent.parent / 'static'))
app.config['SECRET_KEY'] = os.urandom(24).hex()
app.config['MAX_CONTENT_LENGTH'] = APP_CONFIG['max_file_size_mb'] * 1024 * 1024
app.config['UPLOAD_FOLDER'] = str(DATA_DIR / 'uploads')
CORS(app)

# Создаем папку для загрузок
Path(app.config['UPLOAD_FOLDER']).mkdir(parents=True, exist_ok=True)

# Инициализация компонентов
data_loader = DataLoader()
trademark_checker = ComprehensiveTrademarkChecker()
image_checker = ComprehensiveImageChecker()
image_searcher = ComprehensiveImageSearcher()  # Автоматический поиск изображений
risk_evaluator = RiskEvaluator()
export_manager = ExportManager()

# Хранилище сессий (в production использовать БД)
sessions_store: Dict[str, Dict] = {}


def allowed_file(filename: str, allowed_extensions: set) -> bool:
    """Проверка допустимости расширения файла"""
    return '.' in filename and \
           '.' + filename.rsplit('.', 1)[1].lower() in allowed_extensions


@app.route('/')
def index():
    """Главная страница"""
    return render_template('index.html',
                          mktu_classes=MKTU_CLASSES,
                          trademark_resources=TRADEMARK_RESOURCES,
                          image_resources=IMAGE_SEARCH_RESOURCES)


@app.route('/api/upload/excel', methods=['POST'])
def upload_excel():
    """Загрузка Excel/CSV файла с товарами"""
    if 'file' not in request.files:
        return jsonify({'error': 'Файл не найден'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Файл не выбран'}), 400

    if not allowed_file(file.filename, APP_CONFIG['allowed_data_extensions']):
        return jsonify({'error': 'Неподдерживаемый формат файла'}), 400

    try:
        # Сохраняем файл
        filename = secure_filename(file.filename)
        filepath = Path(app.config['UPLOAD_FOLDER']) / filename
        file.save(filepath)

        # Загружаем данные
        items = data_loader.load_from_excel(str(filepath))

        # Создаем сессию
        session = data_loader.create_check_session(items)

        # Сохраняем в хранилище
        sessions_store[session.session_id] = {
            'session': session,
            'assessments': {},
            'created_at': datetime.now().isoformat()
        }

        return jsonify({
            'success': True,
            'session_id': session.session_id,
            'total_items': len(items),
            'items': [
                {
                    'article': item.article,
                    'name': item.name,
                    'category': item.category,
                    'mktu_classes': item.mktu_classes,
                    'image_count': len(item.image_paths),
                    'text_on_product': item.text_on_product,
                    'source_type': item.image_source.source_type if item.image_source else 'unknown'
                }
                for item in items
            ]
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/upload/images', methods=['POST'])
def upload_images():
    """Загрузка изображений из папки или отдельных файлов"""
    if 'files' not in request.files:
        return jsonify({'error': 'Файлы не найдены'}), 400

    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': 'Файлы не выбраны'}), 400

    try:
        items = []
        upload_dir = Path(app.config['UPLOAD_FOLDER']) / str(uuid.uuid4())[:8]
        upload_dir.mkdir(parents=True, exist_ok=True)

        for file in files:
            if file and allowed_file(file.filename, APP_CONFIG['allowed_extensions']):
                filename = secure_filename(file.filename)
                filepath = upload_dir / filename
                file.save(filepath)

                # Создаем ProductItem для каждого изображения
                article = Path(filename).stem
                item = ProductItem(
                    article=article,
                    name=f"Товар {article}",
                    image_paths=[str(filepath)],
                    image_source=ImageSource(source_type='unknown')
                )
                items.append(item)

        if not items:
            return jsonify({'error': 'Нет допустимых изображений'}), 400

        # Создаем сессию
        session = data_loader.create_check_session(items)

        sessions_store[session.session_id] = {
            'session': session,
            'assessments': {},
            'created_at': datetime.now().isoformat()
        }

        return jsonify({
            'success': True,
            'session_id': session.session_id,
            'total_items': len(items),
            'items': [
                {
                    'article': item.article,
                    'name': item.name,
                    'image_path': item.image_paths[0] if item.image_paths else None
                }
                for item in items
            ]
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/check/single', methods=['POST'])
def check_single():
    """Проверка одного товара"""
    data = request.json

    if not data:
        return jsonify({'error': 'Данные не предоставлены'}), 400

    try:
        article = data.get('article', 'MANUAL_' + str(uuid.uuid4())[:8])
        text_to_check = data.get('text', '')
        mktu_classes = data.get('mktu_classes', [])
        image_path = data.get('image_path')

        results = {
            'article': article,
            'trademark_results': [],
            'image_results': None,
            'overall_status': 'green',
            'recommendations': [],
            'manual_check_links': {}
        }

        # Проверка товарных знаков
        if text_to_check:
            tm_results = trademark_checker.check_all(text_to_check, mktu_classes)
            results['trademark_results'] = [
                {
                    'resource': r.resource_name,
                    'status': r.status.value,
                    'exact_match': r.exact_match,
                    'similar_match': r.similar_match,
                    'similarity_score': r.similarity_score,
                    'notes': r.notes,
                    'matches': r.found_matches[:15]  # До 15 совпадений
                }
                for r in tm_results
            ]

            # Ссылки для ручной проверки
            results['manual_check_links'] = trademark_checker.generate_manual_check_links(
                text_to_check, mktu_classes
            )

        # Проверка изображения
        if image_path and os.path.exists(image_path):
            img_results = image_checker.check_image(image_path)
            results['image_results'] = {
                'recognized_texts': [
                    {'text': t.text, 'confidence': t.confidence}
                    for t in img_results.get('recognized_texts', [])
                ],
                'overall_status': img_results.get('overall_status', RiskLevel.GREEN).value,
                'recommendations': img_results.get('recommendations', [])
            }

            # Добавляем ссылки для поиска изображений
            results['manual_check_links'].update(
                img_results.get('manual_check_links', {})
            )

        # Определяем общий статус
        has_red = any(r['status'] == 'red' for r in results['trademark_results'])
        has_yellow = any(r['status'] == 'yellow' for r in results['trademark_results'])

        if results.get('image_results'):
            if results['image_results']['overall_status'] == 'red':
                has_red = True
            elif results['image_results']['overall_status'] == 'yellow':
                has_yellow = True

        if has_red:
            results['overall_status'] = 'red'
        elif has_yellow:
            results['overall_status'] = 'yellow'

        # Сохраняем в историю
        try:
            check_id = save_name_check(
                query_text=text_to_check,
                mktu_classes=mktu_classes,
                overall_status=results['overall_status'],
                results=results['trademark_results'],
                manual_links=results['manual_check_links']
            )
            results['check_id'] = check_id
        except Exception as e:
            print(f"Ошибка сохранения в историю: {e}")

        return jsonify(results)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/check/session/<session_id>', methods=['POST'])
def check_session(session_id):
    """Запуск проверки для всей сессии"""
    if session_id not in sessions_store:
        return jsonify({'error': 'Сессия не найдена'}), 404

    try:
        session_data = sessions_store[session_id]
        session = session_data['session']
        assessments = {}

        for item in session.items:
            # Проверка товарных знаков для текста на товаре
            all_texts = item.text_on_product + item.logos_on_product
            for text in all_texts:
                if text:
                    tm_results = trademark_checker.check_all(text, item.mktu_classes)
                    item.trademark_results.extend(tm_results)

            # Проверка изображений
            for image_path in item.image_paths:
                if os.path.exists(image_path):
                    img_check = image_checker.check_image(image_path)

                    # Добавляем распознанный текст
                    item.recognized_texts.extend(img_check.get('recognized_texts', []))

                    # Проверяем распознанный текст на товарные знаки
                    for text_item in img_check.get('recognized_texts', []):
                        if text_item.text and len(text_item.text) > 2:
                            tm_results = trademark_checker.check_all(
                                text_item.text, item.mktu_classes
                            )
                            item.trademark_results.extend(tm_results)

                    # Результаты поиска изображений
                    item.image_search_results.extend(img_check.get('search_results', []))

                    # Результаты проверки авторских прав
                    if img_check.get('copyright_result'):
                        item.copyright_results.append(img_check['copyright_result'])

            # Оценка риска
            assessment = risk_evaluator.evaluate_product(item)
            assessments[item.article] = assessment

            # Обновляем статус товара
            item.overall_status = assessment.overall_status
            item.status_reason = assessment.summary
            item.recommendations = assessment.recommendations
            item.checked_at = datetime.now()

        # Сохраняем результаты
        session_data['assessments'] = assessments
        session.update_statistics()

        return jsonify({
            'success': True,
            'session_id': session_id,
            'statistics': {
                'total': session.total_items,
                'checked': session.checked_items,
                'red': session.red_count,
                'yellow': session.yellow_count,
                'green': session.green_count
            },
            'results': [
                {
                    'article': item.article,
                    'name': item.name,
                    'status': item.overall_status.value,
                    'risk_score': assessments[item.article].overall_score if item.article in assessments else 0,
                    'summary': item.status_reason,
                    'recommendations': item.recommendations[:3]
                }
                for item in session.items
            ]
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/session/<session_id>')
def get_session(session_id):
    """Получение информации о сессии"""
    if session_id not in sessions_store:
        return jsonify({'error': 'Сессия не найдена'}), 404

    session_data = sessions_store[session_id]
    session = session_data['session']
    assessments = session_data.get('assessments', {})

    return jsonify({
        'session_id': session.session_id,
        'created_at': session.created_at.isoformat(),
        'statistics': {
            'total': session.total_items,
            'red': session.red_count,
            'yellow': session.yellow_count,
            'green': session.green_count
        },
        'items': [
            {
                'article': item.article,
                'name': item.name,
                'status': item.overall_status.value,
                'risk_score': assessments[item.article].overall_score if item.article in assessments else 0,
                'checked': item.checked_at is not None
            }
            for item in session.items
        ]
    })


@app.route('/api/export/<session_id>/<format>')
def export_results(session_id, format):
    """Экспорт результатов"""
    if session_id not in sessions_store:
        return jsonify({'error': 'Сессия не найдена'}), 404

    session_data = sessions_store[session_id]
    session = session_data['session']
    assessments = session_data.get('assessments', {})

    try:
        if format == 'excel':
            filepath = export_manager.export_to_excel(session, assessments)
            return send_file(filepath, as_attachment=True)
        elif format == 'csv':
            filepath = export_manager.export_to_csv(session, assessments)
            return send_file(filepath, as_attachment=True)
        elif format == 'json':
            filepath = export_manager.export_to_json(session, assessments)
            return send_file(filepath, as_attachment=True)
        elif format == 'html':
            filepath = export_manager.export_to_html(session, assessments)
            return send_file(filepath, as_attachment=True)
        else:
            return jsonify({'error': 'Неподдерживаемый формат'}), 400

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/template')
def download_template():
    """Скачивание шаблона Excel"""
    try:
        filepath = TemplateGenerator.create_excel_template()
        return send_file(filepath, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/resources')
def get_resources():
    """Получение списка ресурсов для проверки"""
    return jsonify({
        'trademark_resources': TRADEMARK_RESOURCES,
        'image_resources': IMAGE_SEARCH_RESOURCES,
        'mktu_classes': MKTU_CLASSES
    })


@app.route('/api/check/links', methods=['POST'])
def get_check_links():
    """Получение ссылок для ручной проверки"""
    data = request.json
    text = data.get('text', '')
    mktu_classes = data.get('mktu_classes', [])

    links = trademark_checker.generate_manual_check_links(text, mktu_classes)

    return jsonify({
        'text': text,
        'links': links
    })


@app.route('/api/check/image', methods=['POST'])
def check_image_full():
    """
    Полная проверка изображения:
    1. Загрузка изображения
    2. Распознавание текста (OCR)
    3. Поиск по товарным знакам
    4. Генерация ссылок для обратного поиска
    """
    if 'file' not in request.files:
        return jsonify({'error': 'Файл не найден'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Файл не выбран'}), 400

    if not allowed_file(file.filename, APP_CONFIG['allowed_extensions']):
        return jsonify({'error': 'Неподдерживаемый формат файла'}), 400

    try:
        # Сохраняем файл
        upload_dir = Path(app.config['UPLOAD_FOLDER']) / str(uuid.uuid4())[:8]
        upload_dir.mkdir(parents=True, exist_ok=True)

        from werkzeug.utils import secure_filename
        filename = secure_filename(file.filename)
        filepath = upload_dir / filename
        file.save(filepath)

        # Получаем параметры
        mktu_classes = request.form.getlist('mktu_classes', type=int)
        manual_text = request.form.get('text', '').strip()

        # Результат
        result = {
            'filename': filename,
            'filepath': str(filepath),
            'image_url': f'/uploads/{upload_dir.name}/{filename}',
            'recognized_texts': [],
            'trademark_results': [],
            'image_search_links': {},
            'overall_status': 'green',
            'risk_factors': [],
            'recommendations': [],
            'summary': ''
        }

        # 1. Распознавание текста (OCR)
        try:
            img_check = image_checker.check_image(str(filepath))

            # Получаем распознанный текст (фильтруем по уверенности >= 55%)
            # НЕ показываем ненадёжные распознавания - они дают случайные результаты
            for text_item in img_check.get('recognized_texts', []):
                # Строгий фильтр: только уверенные распознавания (55%+)
                if text_item.confidence < 0.55:
                    continue  # Пропускаем ненадёжные распознавания

                text_clean = text_item.text.strip()
                # Дополнительная фильтрация мусора
                if len(text_clean) < 3:
                    continue
                if not any(c.isalpha() for c in text_clean):
                    continue

                result['recognized_texts'].append({
                    'text': text_clean,
                    'confidence': round(text_item.confidence * 100, 1)
                })

            # Результаты анализа авторских прав
            copyright_result = img_check.get('copyright_result')
            if copyright_result:
                if copyright_result.brand_elements:
                    result['risk_factors'].append({
                        'type': 'brand',
                        'severity': 'red',
                        'message': f"Обнаружены бренды: {', '.join(copyright_result.brand_elements)}"
                    })
                if copyright_result.character_names:
                    result['risk_factors'].append({
                        'type': 'character',
                        'severity': 'red',
                        'message': f"Обнаружены персонажи: {', '.join(copyright_result.character_names)}"
                    })
        except Exception as e:
            result['recommendations'].append(f"Ошибка OCR: {str(e)}")

        # 2. Собираем текст для поиска ТЗ
        texts_to_check = []
        texts_to_check_lower = set()  # Для дедупликации

        # Добавляем ручной текст (имеет приоритет)
        if manual_text:
            texts_to_check.append(manual_text)
            texts_to_check_lower.add(manual_text.lower())

        # Добавляем ТОЛЬКО уверенно распознанный текст (порог уже применён выше)
        for text_item in result['recognized_texts']:
            full_text = text_item['text'].strip()

            # Добавляем полный текст целиком (если > 2 символов)
            if len(full_text) > 2 and full_text.lower() not in texts_to_check_lower:
                texts_to_check.append(full_text)
                texts_to_check_lower.add(full_text.lower())

            # Также добавляем отдельные слова (если > 2 символов)
            words = full_text.split()
            for word in words:
                clean_word = ''.join(c for c in word if c.isalnum())
                if len(clean_word) > 2 and clean_word.lower() not in texts_to_check_lower:
                    texts_to_check.append(clean_word)
                    texts_to_check_lower.add(clean_word.lower())

        # 2.1. Детекция известных брендов по ВСЕМ распознаниям (включая низкоуверенные)
        # Это для предупреждения, но НЕ для автоматического поиска ТЗ
        KNOWN_BRANDS_PATTERNS = {
            'nike': ['nike', 'nke', 'nik', 'nikе', 'niке', 'nіke', 'n1ke', 'nikel'],
            'adidas': ['adidas', 'adldas', 'adіdas', 'ad1das'],
            'puma': ['puma', 'рuma', 'pumа'],
            'gucci': ['gucci', 'guccі', 'guсci'],
            'chanel': ['chanel', 'сhanel', 'chanеl'],
            'louis vuitton': ['vuitton', 'vuіtton', 'lv'],
            'supreme': ['supreme', 'suprеme', 'suprеmе'],
            'champion': ['champion', 'champ1on', 'сhampion', 'champi0n', 'lkpio', 'ckpio', 'chpio'],
        }

        # Собираем ВСЕ распознанные тексты (включая низкоуверенные) для детекции брендов
        all_raw_texts = []
        for text_item in img_check.get('recognized_texts', []):
            if text_item.confidence > 0.15:  # Минимальный порог для детекции брендов
                all_raw_texts.append(text_item.text.lower())
        all_recognized_text = ' '.join(all_raw_texts)

        # Нормализуем текст (заменяем похожие символы)
        normalized_text = all_recognized_text.replace('к', 'k').replace('е', 'e').replace('і', 'i').replace('а', 'a').replace('о', 'o').replace('с', 'c').replace('р', 'p').replace('в', 'b')

        detected_brands = []
        for brand, patterns in KNOWN_BRANDS_PATTERNS.items():
            for pattern in patterns:
                if pattern in normalized_text or pattern in all_recognized_text:
                    detected_brands.append(brand.upper())
                    # НЕ добавляем автоматически в texts_to_check - просто предупреждаем
                    result['risk_factors'].append({
                        'type': 'brand_detected',
                        'severity': 'yellow',  # Жёлтый, т.к. OCR не уверен
                        'message': f"⚠️ Возможно обнаружен бренд: {brand.upper()} (OCR распознал: '{all_recognized_text[:30]}...'). Рекомендуется проверить вручную."
                    })
                    break

        # Если ручной текст не введён И OCR ничего уверенного не нашёл
        if not texts_to_check and not manual_text:
            result['recommendations'].append(
                "⚠️ OCR не смог уверенно распознать текст на изображении. "
                "Если на изображении есть надписи, введите их вручную для проверки товарных знаков."
            )

        # 3. Поиск по товарным знакам (только если есть надёжный текст)
        all_tm_results = []
        checked_texts = []

        # Пропускаем поиск ТЗ если нет надёжного текста
        if not texts_to_check:
            result['recommendations'].append(
                "ℹ️ Проверка товарных знаков не выполнена - нет текста для поиска. "
                "Введите текст вручную, если хотите проверить конкретное название."
            )

        for text in texts_to_check[:5]:  # Максимум 5 проверок
            try:
                tm_results = trademark_checker.check_all(text, mktu_classes)
                checked_texts.append(text)

                for r in tm_results:
                    tm_entry = {
                        'text': text,
                        'resource': r.resource_name,
                        'status': r.status.value,
                        'exact_match': r.exact_match,
                        'similar_match': r.similar_match,
                        'similarity_score': r.similarity_score,
                        'notes': r.notes,
                        'matches': r.found_matches[:5]
                    }
                    all_tm_results.append(tm_entry)

                    # Добавляем факторы риска
                    if r.exact_match:
                        result['risk_factors'].append({
                            'type': 'trademark',
                            'severity': 'red',
                            'message': f"Точное совпадение ТЗ для '{text}': {r.notes}"
                        })
                    elif r.similar_match and r.similarity_score >= 0.8:
                        result['risk_factors'].append({
                            'type': 'trademark',
                            'severity': 'yellow',
                            'message': f"Похожий ТЗ для '{text}' ({r.similarity_score:.0%}): {r.notes}"
                        })

            except Exception as e:
                result['recommendations'].append(f"Ошибка проверки '{text}': {str(e)}")

        result['trademark_results'] = all_tm_results
        result['checked_texts'] = checked_texts

        # 4. Ссылки для ручной проверки ТЗ
        if checked_texts:
            result['trademark_links'] = trademark_checker.generate_manual_check_links(
                checked_texts[0], mktu_classes
            )

        # 5. Автоматический поиск изображений через API
        result['image_search_results'] = []
        result['image_search_links'] = {}

        try:
            # Выполняем автоматический поиск через Serper API
            search_results = image_searcher.search_all(str(filepath), use_api=True)

            for sr in search_results:
                search_result_data = {
                    'resource': sr.resource_name,
                    'url': sr.resource_url,
                    'status': sr.status.value if hasattr(sr.status, 'value') else str(sr.status),
                    'notes': sr.notes,
                    'total_results': sr.total_results,
                    'exact_matches': sr.exact_matches,
                    'similar_images': sr.similar_images[:5] if sr.similar_images else [],
                    'known_sources': sr.known_sources[:5] if sr.known_sources else []
                }
                result['image_search_results'].append(search_result_data)

                # Если найдены совпадения - добавляем в факторы риска
                if sr.status == RiskLevel.RED:
                    result['risk_factors'].append({
                        'type': 'image_search',
                        'severity': 'red',
                        'message': f"🔍 {sr.resource_name}: {sr.notes}"
                    })
                elif sr.status == RiskLevel.YELLOW and sr.total_results > 0:
                    result['risk_factors'].append({
                        'type': 'image_search',
                        'severity': 'yellow',
                        'message': f"🔍 {sr.resource_name}: {sr.notes}"
                    })

        except Exception as e:
            error_msg = str(e)
            # Улучшенное сообщение об ошибке подключения
            if 'Connection' in error_msg or 'timeout' in error_msg.lower():
                result['recommendations'].append(
                    "⚠️ Не удалось выполнить автоматический поиск изображений (ошибка подключения). "
                    "Пожалуйста, используйте ссылки ниже для ручной проверки."
                )
            else:
                result['recommendations'].append(f"⚠️ Ошибка поиска изображений: {error_msg}")

        # Ссылки для ручной проверки (резервный вариант)
        result['image_search_links'] = {
            'yandex': {
                'name': 'Яндекс.Картинки',
                'url': 'https://yandex.ru/images/',
                'instruction': 'Нажмите на иконку камеры и загрузите изображение'
            },
            'google': {
                'name': 'Google Images',
                'url': 'https://images.google.com/',
                'instruction': 'Нажмите на иконку камеры и загрузите изображение'
            },
            'bing': {
                'name': 'Bing Visual Search',
                'url': 'https://www.bing.com/visualsearch',
                'instruction': 'Перетащите изображение для поиска'
            }
        }

        # 6. Определяем общий статус
        has_red = any(f['severity'] == 'red' for f in result['risk_factors'])
        has_yellow = any(f['severity'] == 'yellow' for f in result['risk_factors'])
        has_tm_red = any(r['status'] == 'red' for r in all_tm_results)
        has_tm_yellow = any(r['status'] == 'yellow' for r in all_tm_results)

        if has_red or has_tm_red:
            result['overall_status'] = 'red'
        elif has_yellow or has_tm_yellow:
            result['overall_status'] = 'yellow'
        else:
            result['overall_status'] = 'green'

        # 7. Генерируем рекомендации
        if result['overall_status'] == 'red':
            result['recommendations'].insert(0,
                "⛔ ВНИМАНИЕ: Обнаружены критические совпадения. Использование не рекомендуется без консультации юриста."
            )
        elif result['overall_status'] == 'yellow':
            result['recommendations'].insert(0,
                "⚠️ Требуется дополнительная проверка перед использованием."
            )
        else:
            result['recommendations'].insert(0,
                "✅ Автоматическая проверка не выявила явных проблем."
            )

        if not result['recognized_texts']:
            result['recommendations'].append(
                "Текст на изображении не распознан. Если на изображении есть текст, проверьте его вручную."
            )

        result['recommendations'].append(
            "Рекомендуется выполнить обратный поиск изображения по ссылкам ниже."
        )

        # 8. Сводка
        result['summary'] = {
            'texts_found': len(result['recognized_texts']),
            'texts_checked': len(checked_texts),
            'tm_checks': len(all_tm_results),
            'risk_factors_count': len(result['risk_factors'])
        }

        # 9. Сохраняем в историю
        try:
            check_id = save_image_check(
                filename=filename,
                filepath=str(filepath),
                overall_status=result['overall_status'],
                recognized_texts=result['recognized_texts'],
                trademark_results=result['trademark_results'],
                image_search_results=result.get('image_search_results', []),
                risk_factors=result['risk_factors'],
                recommendations=result['recommendations'],
                summary=result['summary']
            )
            result['check_id'] = check_id
        except Exception as e:
            print(f"Ошибка сохранения в историю: {e}")

        return jsonify(result)

    except Exception as e:
        import traceback
        return jsonify({
            'error': str(e),
            'traceback': traceback.format_exc()
        }), 500


@app.route('/uploads/<path:filename>')
def serve_upload(filename):
    """Отдача загруженных файлов"""
    return send_file(Path(app.config['UPLOAD_FOLDER']) / filename)


# ==================== ИСТОРИЯ ПРОВЕРОК ====================

@app.route('/history')
def history_page():
    """Страница истории проверок"""
    return render_template('history.html',
                          mktu_classes=MKTU_CLASSES,
                          trademark_resources=TRADEMARK_RESOURCES,
                          image_resources=IMAGE_SEARCH_RESOURCES)


@app.route('/api/history/stats')
def get_history_stats():
    """Получить статистику проверок"""
    try:
        stats = get_statistics()
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/history/names')
def get_name_history():
    """Получить историю проверок наименований"""
    try:
        limit = request.args.get('limit', 50, type=int)
        offset = request.args.get('offset', 0, type=int)
        status = request.args.get('status', None)

        checks = get_name_checks(limit=limit, offset=offset, status_filter=status)
        return jsonify({'checks': checks})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/history/images')
def get_image_history():
    """Получить историю проверок изображений"""
    try:
        limit = request.args.get('limit', 50, type=int)
        offset = request.args.get('offset', 0, type=int)
        status = request.args.get('status', None)

        checks = get_image_checks(limit=limit, offset=offset, status_filter=status)
        return jsonify({'checks': checks})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/history/name/<int:check_id>')
def get_name_check_detail(check_id):
    """Получить детали проверки наименования"""
    try:
        check = get_name_check_by_id(check_id)
        if not check:
            return jsonify({'error': 'Проверка не найдена'}), 404
        return jsonify(check)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/history/image/<int:check_id>')
def get_image_check_detail(check_id):
    """Получить детали проверки изображения"""
    try:
        check = get_image_check_by_id(check_id)
        if not check:
            return jsonify({'error': 'Проверка не найдена'}), 404
        return jsonify(check)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/history/delete/<check_type>/<int:check_id>', methods=['DELETE'])
def delete_history_check(check_type, check_id):
    """Удалить проверку из истории"""
    try:
        if check_type not in ['name', 'image']:
            return jsonify({'error': 'Неверный тип проверки'}), 400

        success = delete_check(check_type, check_id)
        if success:
            return jsonify({'success': True})
        return jsonify({'error': 'Проверка не найдена'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/history/clear', methods=['DELETE'])
def clear_all_history():
    """Очистить всю историю"""
    try:
        check_type = request.args.get('type', None)
        deleted = clear_history(check_type)
        return jsonify({'success': True, 'deleted': deleted})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== ЭКСПОРТ ОТЧЁТОВ ====================

@app.route('/api/export/image/<int:check_id>/<format>')
def export_image_report(check_id, format):
    """Экспорт отчёта по проверке изображения"""
    try:
        check = get_image_check_by_id(check_id)
        if not check:
            return jsonify({'error': 'Проверка не найдена'}), 404

        if format == 'excel':
            filepath = export_image_to_excel(check)
            return send_file(filepath, as_attachment=True,
                           download_name=f"report_image_{check_id}.xlsx")
        elif format == 'pdf':
            filepath = export_image_to_pdf(check)
            return send_file(filepath, as_attachment=True,
                           download_name=f"report_image_{check_id}.pdf")
        elif format == 'json':
            return jsonify(check)
        else:
            return jsonify({'error': 'Неподдерживаемый формат'}), 400

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/name/<int:check_id>/<format>')
def export_name_report(check_id, format):
    """Экспорт отчёта по проверке наименования"""
    try:
        check = get_name_check_by_id(check_id)
        if not check:
            return jsonify({'error': 'Проверка не найдена'}), 404

        if format == 'excel':
            filepath = export_name_to_excel(check)
            return send_file(filepath, as_attachment=True,
                           download_name=f"report_name_{check_id}.xlsx")
        elif format == 'pdf':
            filepath = export_name_to_pdf(check)
            return send_file(filepath, as_attachment=True,
                           download_name=f"report_name_{check_id}.pdf")
        elif format == 'json':
            return jsonify(check)
        else:
            return jsonify({'error': 'Неподдерживаемый формат'}), 400

    except Exception as e:
        return jsonify({'error': str(e)}), 500


def export_image_to_excel(check: Dict) -> str:
    """Экспорт проверки изображения в Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчёт проверки"

    # Стили
    header_font = Font(bold=True, size=14)
    status_fills = {
        'red': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        'yellow': PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        'green': PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")
    }

    row = 1

    # Заголовок
    ws.cell(row=row, column=1, value="ОТЧЁТ О ПРОВЕРКЕ ИЗОБРАЖЕНИЯ").font = Font(bold=True, size=16)
    row += 2

    # Основная информация
    ws.cell(row=row, column=1, value="Файл:").font = header_font
    ws.cell(row=row, column=2, value=check.get('filename', '-'))
    row += 1

    ws.cell(row=row, column=1, value="Дата проверки:").font = header_font
    ws.cell(row=row, column=2, value=check.get('created_at', '-'))
    row += 1

    ws.cell(row=row, column=1, value="Статус:").font = header_font
    status = check.get('overall_status', 'green')
    status_text = {'red': 'ЗАПРЕЩЕНО', 'yellow': 'ТРЕБУЕТ ПРОВЕРКИ', 'green': 'РАЗРЕШЕНО'}.get(status, status)
    cell = ws.cell(row=row, column=2, value=status_text)
    cell.fill = status_fills.get(status, status_fills['green'])
    row += 2

    # Распознанные тексты
    ws.cell(row=row, column=1, value="РАСПОЗНАННЫЕ ТЕКСТЫ").font = header_font
    row += 1
    texts = check.get('recognized_texts', [])
    if texts:
        for t in texts:
            ws.cell(row=row, column=1, value=t.get('text', '-'))
            ws.cell(row=row, column=2, value=f"{t.get('confidence', 0)}%")
            row += 1
    else:
        ws.cell(row=row, column=1, value="Текст не распознан")
        row += 1
    row += 1

    # Факторы риска
    ws.cell(row=row, column=1, value="ФАКТОРЫ РИСКА").font = header_font
    row += 1
    risks = check.get('risk_factors', [])
    if risks:
        for r in risks:
            ws.cell(row=row, column=1, value=r.get('message', '-'))
            row += 1
    else:
        ws.cell(row=row, column=1, value="Не обнаружено")
        row += 1
    row += 1

    # Рекомендации
    ws.cell(row=row, column=1, value="РЕКОМЕНДАЦИИ").font = header_font
    row += 1
    recs = check.get('recommendations', [])
    for r in recs:
        ws.cell(row=row, column=1, value=r)
        row += 1

    # Автоширина колонок
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 50

    # Сохранение
    output_path = OUTPUT_DIR / f"report_image_{check.get('id', 'unknown')}.xlsx"
    wb.save(str(output_path))
    return str(output_path)


def export_image_to_pdf(check: Dict) -> str:
    """Экспорт проверки изображения в PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # Регистрируем шрифт с поддержкой кириллицы
    try:
        pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        font_name = 'DejaVu'
    except:
        font_name = 'Helvetica'

    output_path = OUTPUT_DIR / f"report_image_{check.get('id', 'unknown')}.pdf"
    doc = SimpleDocTemplate(str(output_path), pagesize=A4,
                           rightMargin=2*cm, leftMargin=2*cm,
                           topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='RuTitle', fontName=font_name, fontSize=18, spaceAfter=20))
    styles.add(ParagraphStyle(name='RuHeading', fontName=font_name, fontSize=14, spaceAfter=10, spaceBefore=15))
    styles.add(ParagraphStyle(name='RuNormal', fontName=font_name, fontSize=11, spaceAfter=5))

    story = []

    # Заголовок
    story.append(Paragraph("OTCHET O PROVERKE IZOBRAZHENIYA", styles['RuTitle']))
    story.append(Spacer(1, 0.5*cm))

    # Статус
    status = check.get('overall_status', 'green')
    status_text = {'red': 'ZAPRESHCHENO', 'yellow': 'TREBUET PROVERKI', 'green': 'RAZRESHENO'}.get(status, status)
    status_color = {'red': colors.red, 'yellow': colors.yellow, 'green': colors.green}.get(status, colors.green)

    # Основная информация
    data = [
        ['Fayl:', check.get('filename', '-')],
        ['Data:', check.get('created_at', '-')],
        ['Status:', status_text],
    ]
    t = Table(data, colWidths=[4*cm, 12*cm])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('FONTNAME', (0, 0), (0, -1), font_name),
        ('BACKGROUND', (1, 2), (1, 2), status_color),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    # Распознанные тексты
    story.append(Paragraph("Raspoznannye teksty:", styles['RuHeading']))
    texts = check.get('recognized_texts', [])
    if texts:
        for t in texts:
            story.append(Paragraph(f"• {t.get('text', '-')} ({t.get('confidence', 0)}%)", styles['RuNormal']))
    else:
        story.append(Paragraph("Tekst ne raspoznan", styles['RuNormal']))

    # Факторы риска
    story.append(Paragraph("Faktory riska:", styles['RuHeading']))
    risks = check.get('risk_factors', [])
    if risks:
        for r in risks:
            story.append(Paragraph(f"• {r.get('message', '-')}", styles['RuNormal']))
    else:
        story.append(Paragraph("Ne obnaruzheno", styles['RuNormal']))

    # Рекомендации
    story.append(Paragraph("Rekomendatsii:", styles['RuHeading']))
    recs = check.get('recommendations', [])
    for r in recs:
        # Убираем эмодзи для PDF
        r_clean = r.replace('⛔', '[!]').replace('⚠️', '[!]').replace('✅', '[OK]')
        story.append(Paragraph(f"• {r_clean}", styles['RuNormal']))

    doc.build(story)
    return str(output_path)


def export_name_to_excel(check: Dict) -> str:
    """Экспорт проверки наименования в Excel"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчёт проверки"

    header_font = Font(bold=True, size=14)
    status_fills = {
        'red': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        'yellow': PatternFill(start_color="FFE66D", end_color="FFE66D", fill_type="solid"),
        'green': PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")
    }

    row = 1

    ws.cell(row=row, column=1, value="ОТЧЁТ О ПРОВЕРКЕ НАИМЕНОВАНИЯ").font = Font(bold=True, size=16)
    row += 2

    ws.cell(row=row, column=1, value="Текст запроса:").font = header_font
    ws.cell(row=row, column=2, value=check.get('query_text', '-'))
    row += 1

    ws.cell(row=row, column=1, value="Классы МКТУ:").font = header_font
    ws.cell(row=row, column=2, value=', '.join(map(str, check.get('mktu_classes', []))) or '-')
    row += 1

    ws.cell(row=row, column=1, value="Дата проверки:").font = header_font
    ws.cell(row=row, column=2, value=check.get('created_at', '-'))
    row += 1

    ws.cell(row=row, column=1, value="Статус:").font = header_font
    status = check.get('overall_status', 'green')
    status_text = {'red': 'ЗАПРЕЩЕНО', 'yellow': 'ТРЕБУЕТ ПРОВЕРКИ', 'green': 'РАЗРЕШЕНО'}.get(status, status)
    cell = ws.cell(row=row, column=2, value=status_text)
    cell.fill = status_fills.get(status, status_fills['green'])
    row += 2

    # Результаты проверки
    ws.cell(row=row, column=1, value="РЕЗУЛЬТАТЫ ПРОВЕРКИ").font = header_font
    row += 1
    results = check.get('results', [])
    for r in results:
        ws.cell(row=row, column=1, value=r.get('resource', '-'))
        ws.cell(row=row, column=2, value=r.get('notes', '-'))
        row += 1
    row += 1

    # Ссылки для проверки
    ws.cell(row=row, column=1, value="ССЫЛКИ ДЛЯ РУЧНОЙ ПРОВЕРКИ").font = header_font
    row += 1
    links = check.get('manual_links', {})
    for name, url in links.items():
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=url)
        row += 1

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 60

    output_path = OUTPUT_DIR / f"report_name_{check.get('id', 'unknown')}.xlsx"
    wb.save(str(output_path))
    return str(output_path)


def export_name_to_pdf(check: Dict) -> str:
    """Экспорт проверки наименования в PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    try:
        pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        font_name = 'DejaVu'
    except:
        font_name = 'Helvetica'

    output_path = OUTPUT_DIR / f"report_name_{check.get('id', 'unknown')}.pdf"
    doc = SimpleDocTemplate(str(output_path), pagesize=A4,
                           rightMargin=2*cm, leftMargin=2*cm,
                           topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='RuTitle', fontName=font_name, fontSize=18, spaceAfter=20))
    styles.add(ParagraphStyle(name='RuHeading', fontName=font_name, fontSize=14, spaceAfter=10, spaceBefore=15))
    styles.add(ParagraphStyle(name='RuNormal', fontName=font_name, fontSize=11, spaceAfter=5))

    story = []

    story.append(Paragraph("OTCHET O PROVERKE NAIMENOVANIYA", styles['RuTitle']))
    story.append(Spacer(1, 0.5*cm))

    status = check.get('overall_status', 'green')
    status_text = {'red': 'ZAPRESHCHENO', 'yellow': 'TREBUET PROVERKI', 'green': 'RAZRESHENO'}.get(status, status)
    status_color = {'red': colors.red, 'yellow': colors.yellow, 'green': colors.green}.get(status, colors.green)

    data = [
        ['Tekst:', check.get('query_text', '-')],
        ['Klassy MKTU:', ', '.join(map(str, check.get('mktu_classes', []))) or '-'],
        ['Data:', check.get('created_at', '-')],
        ['Status:', status_text],
    ]
    t = Table(data, colWidths=[4*cm, 12*cm])
    t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), font_name),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BACKGROUND', (1, 3), (1, 3), status_color),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*cm))

    story.append(Paragraph("Rezultaty proverki:", styles['RuHeading']))
    results = check.get('results', [])
    for r in results:
        story.append(Paragraph(f"• {r.get('resource', '-')}: {r.get('notes', '-')}", styles['RuNormal']))

    story.append(Paragraph("Ssylki dlya proverki:", styles['RuHeading']))
    links = check.get('manual_links', {})
    for name, url in links.items():
        story.append(Paragraph(f"• {name}: {url}", styles['RuNormal']))

    doc.build(story)
    return str(output_path)


if __name__ == '__main__':
    print("=" * 60)
    print("Система проверки интеллектуальной собственности")
    print("=" * 60)
    print(f"Запуск веб-сервера на http://localhost:{APP_CONFIG['port']}")
    print("Для остановки нажмите Ctrl+C")
    print("=" * 60)

    app.run(
        host=APP_CONFIG['host'],
        port=APP_CONFIG['port'],
        debug=APP_CONFIG['debug']
    )
